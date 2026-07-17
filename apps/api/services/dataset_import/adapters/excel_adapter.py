"""
AstroOS — Generic Excel Adapter

Reads any Excel (.xlsx/.xls) file and yields RawRecord objects.
Handles multiple sheets, header detection, and basic type inference.
"""

from __future__ import annotations

import os
from typing import Any, Dict, Iterator, List

import openpyxl

from apps.api.services.dataset_import.adapter_base import ColumnDefinition, RawRecord, SourceAdapter


class ExcelAdapter(SourceAdapter):
    """Generic Excel adapter for .xlsx files."""

    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    def __init__(self, sheet_name: str = None):
        """
        Args:
            sheet_name: Specific sheet to read. None reads the first sheet.
        """
        self._sheet_name = sheet_name

    def read(self, file_path: str) -> Iterator[RawRecord]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[self._sheet_name] if self._sheet_name and self._sheet_name in wb.sheetnames else wb.worksheets[0]

            rows = list(ws.iter_rows())
            if not rows:
                return

            headers = [str(cell.value).strip() if cell.value else f"col_{i}"
                       for i, cell in enumerate(rows[0])]

            for idx, row in enumerate(rows[1:], start=1):
                data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        data[headers[i]] = cell.value
                yield RawRecord(
                    data=data,
                    source_index=idx,
                    source_file=os.path.basename(file_path),
                    source_sheet=ws.title,
                )
        finally:
            wb.close()

    def get_source_metadata(self, file_path: str) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[self._sheet_name] if self._sheet_name and self._sheet_name in wb.sheetnames else wb.worksheets[0]
            record_count = max(0, ws.max_row - 1) if ws.max_row else 0
            headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            file_size = os.path.getsize(file_path)

            return {
                "source_file": os.path.basename(file_path),
                "source_format": "xlsx",
                "record_count": record_count,
                "source_columns": headers,
                "sheet_name": ws.title,
                "file_size_bytes": file_size,
                "total_sheets": len(wb.sheetnames),
            }
        finally:
            wb.close()

    def get_column_definitions(self, file_path: str) -> List[ColumnDefinition]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[self._sheet_name] if self._sheet_name and self._sheet_name in wb.sheetnames else wb.worksheets[0]

            rows = list(ws.iter_rows())
            if len(rows) < 2:
                return []

            headers = [str(cell.value).strip() if cell.value else f"col_{i}"
                       for i, cell in enumerate(rows[0])]

            # Sample first 10 data rows for type inference
            sample_rows = rows[1:11]
            columns = []

            for i, header in enumerate(headers):
                values = [row[i].value for row in sample_rows if i < len(row)]
                non_null = [v for v in values if v is not None]

                # Infer type
                if not non_null:
                    data_type = "string"
                    nullable = True
                elif all(isinstance(v, (int, float)) for v in non_null):
                    data_type = "number"
                    nullable = any(v is None for v in values)
                else:
                    data_type = "string"
                    nullable = any(v is None for v in values)

                columns.append(ColumnDefinition(
                    name=header,
                    data_type=data_type,
                    nullable=nullable,
                    example=non_null[0] if non_null else None,
                ))

            return columns
        finally:
            wb.close()

    def get_adapter_name(self) -> str:
        return "Excel"

    def get_supported_formats(self) -> List[str]:
        return self.SUPPORTED_EXTENSIONS
