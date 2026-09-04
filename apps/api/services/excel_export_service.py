"""
AstroOS — Excel Export Service for Kurma Chakra & SBC Vedha Map.

Builds structured, formula-linked Excel workbooks containing:
1. 'Kurma_Display' worksheet with 9-directional Kurma Chakra overview and
   Pada112 Z-AC layout (columns Z, AA, AB, AC).
2. 'tblVedhaMap_Sheet' with an Excel Table (ListObject) named 'tblVedhaMap'
   containing the 112-pada Sarvatobhadra Chakra vedha cross-reference.
3. Structured XLOOKUP / INDEX(MATCH()) formulas linking display ranges to tblVedhaMap.
"""

from __future__ import annotations

import csv
import io
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from apps.api.domain.mundane import KurmaChakraState, KurmaSectorStatus
from apps.api.services.ephemeris_wrapper import EphemerisWrapper
from apps.api.services.kurma_chakra_engine import KurmaChakraEngine


class ExcelExportService:
    """Generates production-grade Excel workbooks and data payloads for Office.js and direct download."""

    def __init__(self, wrapper: Optional[EphemerisWrapper] = None):
        self._wrapper = wrapper or EphemerisWrapper(ephemeris_path="data/ephemeris")
        self._kurma_engine = KurmaChakraEngine(self._wrapper)
        self._vedha_csv_path = Path(__file__).resolve().parents[3] / "docs" / "reference-data" / "sbc_vedha_map_112.csv"

    def get_vedha_map_rows(self) -> List[Dict[str, Any]]:
        """Reads and returns the 112-pada SBC vedha map."""
        if not self._vedha_csv_path.exists():
            raise FileNotFoundError(f"SBC Vedha map CSV not found at {self._vedha_csv_path}")

        rows = []
        with open(self._vedha_csv_path, mode="r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append({
                    "pada112": int(r["pada112"]),
                    "position": r["position"].strip(),
                    "left": r["left"].strip(),
                    "front": r["front"].strip(),
                    "right": r["right"].strip(),
                })
        return rows

    def generate_kurma_workbook_bytes(
        self,
        dt: Optional[datetime] = None,
        ayanamsa: str = "lahiri",
    ) -> bytes:
        """Generates the full Kurma Chakra & tblVedhaMap Excel workbook in memory."""
        dt = dt or datetime.now(timezone.utc)
        state: KurmaChakraState = self._kurma_engine.evaluate_state(dt, ayanamsa)
        vedha_rows = self.get_vedha_map_rows()

        wb = openpyxl.Workbook()

        # Sheet 1: Kurma_Display
        ws_display = wb.active
        ws_display.title = "Kurma_Display"
        ws_display.views.sheetView[0].showGridLines = True

        # Sheet 2: tblVedhaMap_Sheet (Data Source Table)
        ws_table = wb.create_sheet(title="tblVedhaMap_Sheet")
        ws_table.views.sheetView[0].showGridLines = True

        self._populate_vedha_table_sheet(ws_table, vedha_rows)
        self._populate_kurma_display_sheet(ws_display, state, dt, ayanamsa, len(vedha_rows))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _populate_vedha_table_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, rows: List[Dict[str, Any]]) -> None:
        """Populates the tblVedhaMap sheet and defines the official ListObject Table."""
        headers = ["pada112", "position", "left", "front", "right"]
        ws.append(headers)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in rows:
            ws.append([r["pada112"], r["position"], r["left"], r["front"], r["right"]])

        num_rows = len(rows) + 1
        table = Table(displayName="tblVedhaMap", ref=f"A1:E{num_rows}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def _populate_kurma_display_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        state: KurmaChakraState,
        dt: datetime,
        ayanamsa: str,
        total_padas: int,
    ) -> None:
        """Fills Kurma_Display with header info, 9-directional status, and Z-AC Pada matrix."""
        # Styling definitions
        title_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
        subtitle_font = Font(name="Calibri", size=10, italic=True, color="64748B")
        section_font = Font(name="Calibri", size=12, bold=True, color="0284C7")
        header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        afflicted_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        safe_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

        # 1. Main Header
        ws["A1"] = "ASTROOS — KURMA CHAKRA & SARVATOBHADRA VEDHA CONSOLE"
        ws["A1"].font = title_font
        ws["A2"] = f"Evaluated Timestamp: {dt.isoformat()} | Ayanamsa: {ayanamsa.capitalize()} | Summary: {state.summary}"
        ws["A2"].font = subtitle_font

        # 2. Kurma Chakra 9-Directional Summary Table
        ws["A4"] = "9-DIRECTIONAL GEOPOLITICAL VULNERABILITY SUMMARY"
        ws["A4"].font = section_font

        summary_headers = [
            "Direction",
            "Nakshatras",
            "Traditional Regions",
            "Transiting Malefics",
            "Transiting Benefics",
            "Status",
            "Severity",
            "Risk Summary",
        ]
        for col_idx, h in enumerate(summary_headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        ws.row_dimensions[5].height = 24

        curr_row = 6
        for s in state.sectors:
            ws.cell(row=curr_row, column=1, value=s.direction.value.capitalize()).alignment = Alignment(horizontal="center")
            ws.cell(row=curr_row, column=2, value=", ".join(s.nakshatras))
            ws.cell(row=curr_row, column=3, value=", ".join(s.traditional_regions))
            ws.cell(row=curr_row, column=4, value=", ".join(s.transiting_malefics) or "None").alignment = Alignment(horizontal="center")
            ws.cell(row=curr_row, column=5, value=", ".join(s.transiting_benefics) or "None").alignment = Alignment(horizontal="center")
            
            status_cell = ws.cell(row=curr_row, column=6, value="AFFLICTED" if s.is_afflicted else "BENIGN")
            status_cell.alignment = Alignment(horizontal="center")
            status_cell.fill = afflicted_fill if s.is_afflicted else safe_fill
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="991B1B" if s.is_afflicted else "166534")

            ws.cell(row=curr_row, column=7, value=str(s.severity)).alignment = Alignment(horizontal="center")
            ws.cell(row=curr_row, column=8, value=s.risk_summary)

            for c in range(1, 9):
                ws.cell(row=curr_row, column=c).border = thin_border

            ws.row_dimensions[curr_row].height = 20
            curr_row += 1

        # 3. Pada 1-112 Matrix with Z-AC (Cols 26..29) Layout starting at row 26
        ws["Y24"] = "PADA 1-112 SBC VEDHA MATRIX (Z-AC MAPPING)"
        ws["Y24"].font = section_font

        # Column letters: Y=25 (pada112 key), Z=26 (position), AA=27 (left), AB=28 (front), AC=29 (right)
        matrix_headers = [
            (25, "Pada #"),
            (26, "Position (Z)"),
            (27, "Left Vedha (AA)"),
            (28, "Front Vedha (AB)"),
            (29, "Right Vedha (AC)"),
        ]

        for col_idx, h in matrix_headers:
            cell = ws.cell(row=25, column=col_idx, value=h)
            cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.row_dimensions[25].height = 22

        # Populate rows 26 onwards with key in Col Y and INDEX/MATCH formulas in Z, AA, AB, AC
        for pada_idx in range(1, min(total_padas + 1, 113)):
            r_idx = 25 + pada_idx
            # Key cell (Col Y / 25)
            ws.cell(row=r_idx, column=25, value=pada_idx).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=25).border = thin_border

            # Position (Col Z / 26): formula =INDEX(tblVedhaMap[position], MATCH(Y26, tblVedhaMap[pada112], 0))
            pos_cell = ws.cell(
                row=r_idx,
                column=26,
                value=f'=INDEX(tblVedhaMap[position], MATCH(Y{r_idx}, tblVedhaMap[pada112], 0))',
            )
            pos_cell.border = thin_border

            # Left Vedha (Col AA / 27)
            left_cell = ws.cell(
                row=r_idx,
                column=27,
                value=f'=INDEX(tblVedhaMap[left], MATCH(Y{r_idx}, tblVedhaMap[pada112], 0))',
            )
            left_cell.border = thin_border

            # Front Vedha (Col AB / 28)
            front_cell = ws.cell(
                row=r_idx,
                column=28,
                value=f'=INDEX(tblVedhaMap[front], MATCH(Y{r_idx}, tblVedhaMap[pada112], 0))',
            )
            front_cell.border = thin_border

            # Right Vedha (Col AC / 29)
            right_cell = ws.cell(
                row=r_idx,
                column=29,
                value=f'=INDEX(tblVedhaMap[right], MATCH(Y{r_idx}, tblVedhaMap[pada112], 0))',
            )
            right_cell.border = thin_border

            ws.row_dimensions[r_idx].height = 18

        # Adjust column widths for A-H and Y-AC
        col_widths = {
            "A": 16,
            "B": 30,
            "C": 30,
            "D": 22,
            "E": 22,
            "F": 16,
            "G": 12,
            "H": 45,
            "Y": 10,
            "Z": 18,
            "AA": 22,
            "AB": 22,
            "AC": 22,
        }
        for col_name, width in col_widths.items():
            ws.column_dimensions[col_name].width = width
