"""
Tests for ExcelExportService & Kurma / SBC Vedha Map Excel generation.
"""

import io
from datetime import datetime, timezone
import openpyxl
import pytest

from apps.api.services.excel_export_service import ExcelExportService


def test_vedha_map_rows_loading():
    service = ExcelExportService()
    rows = service.get_vedha_map_rows()
    assert len(rows) == 112
    assert rows[0]["pada112"] == 1
    assert rows[0]["position"] == "Krittika-1"
    assert rows[0]["left"] == "Visakha-4"
    assert rows[0]["front"] == "Sravana-4"
    assert rows[0]["right"] == "Bharani-4"


def test_generate_kurma_workbook_bytes():
    service = ExcelExportService()
    dt = datetime(2026, 3, 19, 6, 0, 0, tzinfo=timezone.utc)
    wb_bytes = service.generate_kurma_workbook_bytes(dt=dt, ayanamsa="lahiri")

    assert isinstance(wb_bytes, bytes)
    assert len(wb_bytes) > 0

    # Inspect the generated workbook with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
    assert "Kurma_Display" in wb.sheetnames
    assert "tblVedhaMap_Sheet" in wb.sheetnames

    # Check Table (ListObject) on tblVedhaMap_Sheet
    ws_table = wb["tblVedhaMap_Sheet"]
    assert len(ws_table.tables) == 1
    assert "tblVedhaMap" in ws_table.tables
    table = ws_table.tables["tblVedhaMap"]
    assert table.ref == "A1:E113"

    # Check Kurma_Display contents & formulas
    ws_display = wb["Kurma_Display"]
    assert "ASTROOS — KURMA CHAKRA" in str(ws_display["A1"].value)
    assert ws_display["A5"].value == "Direction"

    # Verify Pada matrix starting at row 26
    # Y26 is pada 1, Z26 is formula referencing tblVedhaMap[position]
    assert ws_display["Y26"].value == 1
    assert "=INDEX(tblVedhaMap[position]" in str(ws_display["Z26"].value)
    assert "=INDEX(tblVedhaMap[left]" in str(ws_display["AA26"].value)
    assert "=INDEX(tblVedhaMap[front]" in str(ws_display["AB26"].value)
    assert "=INDEX(tblVedhaMap[right]" in str(ws_display["AC26"].value)


def test_excel_router_endpoints():
    from fastapi.testclient import TestClient
    from apps.api.main import create_app
    from apps.api.dependencies import require_authenticated

    app = create_app()
    # Bypass auth dependency for unit test
    app.dependency_overrides[require_authenticated] = lambda: {"user_id": "test_user"}
    client = TestClient(app)

    # 1. Test vedha-map-data
    res = client.get("/api/v1/excel/vedha-map-data")
    assert res.status_code == 200
    data = res.json()
    assert data["count"] == 112
    assert "rows" in data
    assert data["rows"][0]["position"] == "Krittika-1"

    # 2. Test kurma-display-payload
    res_payload = client.get("/api/v1/excel/kurma-display-payload")
    assert res_payload.status_code == 200
    pdata = res_payload.json()
    assert len(pdata["summary_values"]) == 9
    assert pdata["summary_range"] == "A6:H14"

    # 3. Test export-kurma xlsx stream
    res_stream = client.get("/api/v1/excel/export-kurma")
    assert res_stream.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in res_stream.headers["content-type"]
    assert len(res_stream.content) > 10000

