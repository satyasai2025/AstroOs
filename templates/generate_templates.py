"""Generate AstroOS cohort import templates in CSV and XLSX formats."""

import csv
import os
from pathlib import Path

TEMPLATE_DIR = Path(__file__).parent


def _capitalize_name(name: str) -> str:
    """Capitalize each word in a name/place string."""
    return " ".join(word.capitalize() for word in name.strip().split())


HEADERS = [
    "_record_id",
    "subject_first_name",
    "subject_last_name",
    "subject_gender",
    "birth_place",
    "birth_country",
    "birth_latitude",
    "birth_longitude",
    "birth_day",
    "birth_month",
    "birth_year",
    "birth_hour",
    "birth_minute",
    "calendar_type",
    "julian_day_ut",
    "birth_time_precision",
    "birth_datetime_utc",
    "source_type",
    "confidence_tier",
]

EXAMPLE_ROWS = [
    [
        "ASTRO-REC-COHORT-000001",
        _capitalize_name("Ada"),
        _capitalize_name("Lovelace"),
        "F",
        _capitalize_name("London"),
        "England",
        51.5072,
        -0.1275,
        10,
        "Dec",
        1815,
        15,
        0,
        "g",
        2379667.5,
        "AA",
        "1815-12-10T15:00:00+00:00",
        "biography",
        "verified",
    ],
    [
        "ASTRO-REC-COHORT-000002",
        _capitalize_name("Nikola"),
        _capitalize_name("Tesla"),
        "M",
        _capitalize_name("Smiljan"),
        "Croatia",
        44.5,
        15.3,
        10,
        "Jul",
        1856,
        0,
        0,
        "g",
        2398029.5,
        "AA",
        "1856-07-10T00:00:00+00:00",
        "biography",
        "verified",
    ],
    [
        "ASTRO-REC-COHORT-000003",
        _capitalize_name("Marie"),
        _capitalize_name("Curie"),
        "F",
        _capitalize_name("Warsaw"),
        "Poland",
        52.2297,
        21.0122,
        7,
        "Nov",
        1867,
        0,
        0,
        "g",
        2404043.5,
        "A",
        "1867-11-07T00:00:00+00:00",
        "biography",
        "estimated",
    ],
]


def generate_csv(path: str):
    """Generate a CSV import template."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(EXAMPLE_ROWS)
    print(f"  [OK] CSV template: {path}")


def generate_xlsx(path: str):
    """Generate an XLSX import template with headers and example data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cohort Import"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Example data rows
    for row_idx, row_data in enumerate(EXAMPLE_ROWS, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Column widths
    widths = {
        1: 24,   # _record_id
        2: 20,   # first_name
        3: 20,   # last_name
        4: 14,   # gender
        5: 18,   # birth_place
        6: 14,   # birth_country
        7: 14,   # birth_latitude
        8: 14,   # birth_longitude
        9: 10,   # birth_day
        10: 10,  # birth_month
        11: 10,  # birth_year
        12: 10,  # birth_hour
        13: 10,  # birth_minute
        14: 14,  # calendar_type
        15: 14,  # julian_day_ut
        16: 18,  # birth_time_precision
        17: 28,  # birth_datetime_utc
        18: 14,  # source_type
        19: 16,  # confidence_tier
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(path)
    print(f"  [OK] XLSX template: {path}")


def main():
    print("Generating AstroOS import templates...")
    generate_csv(str(TEMPLATE_DIR / "astrosos-cohort-import-template.csv"))
    generate_xlsx(str(TEMPLATE_DIR / "astrosos-cohort-import-template.xlsx"))
    print("Done.")


if __name__ == "__main__":
    main()